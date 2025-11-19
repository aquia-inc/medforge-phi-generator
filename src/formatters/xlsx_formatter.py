"""
XLSX (Excel) formatter for PHI documents
Creates spreadsheets with patient data and de-identified templates
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class XLSXFormatter:
    """Creates Excel spreadsheets with PHI content"""

    def __init__(self, output_dir='output'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def create_lab_results_spreadsheet(self, patient, provider, facility, lab_data, filename):
        """Create lab results spreadsheet (PHI Positive)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Lab Results"

        # Header styling
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Facility header
        ws['A1'] = facility['name'].upper()
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"{facility['address']['street']}, {facility['address']['city']}, {facility['address']['state']}"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Document title
        ws['A4'] = "LABORATORY RESULTS"
        ws['A4'].font = Font(bold=True, size=13)
        ws.merge_cells('A4:F4')
        ws['A4'].alignment = Alignment(horizontal='center')

        # Patient Information
        row = 6
        ws[f'A{row}'] = "PATIENT INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")

        patient_info = [
            ("Patient Name:", f"{patient['last_name']}, {patient['first_name']}"),
            ("Date of Birth:", patient['dob'].strftime('%m/%d/%Y')),
            ("Age:", str(patient['age'])),
            ("MRN:", patient['mrn']),
            ("Address:", f"{patient['address']['street']}, {patient['address']['city']}, {patient['address']['state']} {patient['address']['zip']}"),
            ("Phone:", patient['phone'])
        ]

        row += 1
        for label, value in patient_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Test Information
        row += 1
        ws[f'A{row}'] = "TEST INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")

        test_info = [
            ("Collection Date:", lab_data['test_date'].strftime('%m/%d/%Y')),
            ("Report Date:", datetime.now().strftime('%m/%d/%Y')),
            ("Ordering Provider:", f"{provider['first_name']} {provider['last_name']}, {provider['title']}")
        ]

        row += 1
        for label, value in test_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Lab Results Table
        row += 2
        ws[f'A{row}'] = "LABORATORY RESULTS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        headers = ['Test Name', 'Result', 'Unit', 'Reference Range', 'Flag']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Results data
        for result in lab_data['results']:
            row += 1
            ws.cell(row=row, column=1, value=result['test'])
            ws.cell(row=row, column=2, value=result['value'])
            ws.cell(row=row, column=3, value=result['unit'])
            ws.cell(row=row, column=4, value=result['reference_range'])
            flag_cell = ws.cell(row=row, column=5, value=result.get('flag', ''))

            # Highlight abnormal results
            if result.get('flag'):
                flag_cell.font = Font(color="FF0000", bold=True)

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 8

        # Footer
        row += 2
        ws[f'A{row}'] = "CONFIDENTIAL - Protected Health Information"
        ws[f'A{row}'].font = Font(size=9, italic=True)
        ws.merge_cells(f'A{row}:E{row}')

        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def create_patient_roster(self, patients, facility, filename):
        """Create patient roster spreadsheet (PHI Negative - Aggregated/De-identified)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Patient Statistics"

        # Header
        ws['A1'] = f"{facility['name']} - Patient Demographics Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Report Date: {datetime.now().strftime('%m/%d/%Y')}"
        ws.merge_cells('A2:D2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Column headers
        row = 4
        headers = ['Age Group', 'Count', 'Percentage', 'Primary Diagnosis Categories']
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Aggregate data (no individual patient identifiers)
        age_groups = [
            ("18-30", len([p for p in patients if 18 <= p['age'] <= 30])),
            ("31-50", len([p for p in patients if 31 <= p['age'] <= 50])),
            ("51-70", len([p for p in patients if 51 <= p['age'] <= 70])),
            ("71+", len([p for p in patients if p['age'] > 70]))
        ]

        total = len(patients)
        row += 1
        for age_group, count in age_groups:
            ws.cell(row=row, column=1, value=age_group)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count/total*100):.1f}%")
            ws.cell(row=row, column=4, value="Diabetes, Hypertension, Hyperlipidemia")
            row += 1

        # Note
        row += 2
        ws[f'A{row}'] = "Note: This report contains aggregated, de-identified data only. No individual patient information is included."
        ws[f'A{row}'].font = Font(size=9, italic=True)
        ws.merge_cells(f'A{row}:D{row}')

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def create_billing_summary(self, facility, filename):
        """Create generic billing summary (PHI Negative - No Patient Data)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Summary"

        # Header
        ws['A1'] = f"{facility['name']} - Monthly Billing Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = "Period: December 2024"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Column headers
        row = 4
        headers = ['Service Category', 'Procedure Count', 'Average Charge', 'Total Charges', 'Payment Rate']
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Generic billing data (no patient identifiers)
        billing_data = [
            ("Office Visits", 245, 150.00, 36750.00, "92%"),
            ("Lab Services", 189, 85.00, 16065.00, "88%"),
            ("Imaging Studies", 67, 425.00, 28475.00, "85%"),
            ("Procedures", 34, 650.00, 22100.00, "90%"),
            ("Consultations", 56, 225.00, 12600.00, "93%")
        ]

        row += 1
        for category, count, avg, total, rate in billing_data:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"${avg:.2f}")
            ws.cell(row=row, column=4, value=f"${total:.2f}")
            ws.cell(row=row, column=5, value=rate)
            row += 1

        # Totals
        ws.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=row, column=2, value=sum([d[1] for d in billing_data])).font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"${sum([d[3] for d in billing_data]):.2f}").font = Font(bold=True)

        # Note
        row += 2
        ws[f'A{row}'] = "Note: Summary data only. No individual patient information is included."
        ws[f'A{row}'].font = Font(size=9, italic=True)
        ws.merge_cells(f'A{row}:E{row}')

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
